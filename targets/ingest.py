"""Lead ingestion from an Excel sheet.

Parses an .xlsx (flexible column names) and creates leads in the caller's
workspace with whatever context the sheet carries — name, phone, config/budget,
status, score, and any notes/interactions become the lead's starting memory.
"""

import io
from typing import Dict, Any, List

import openpyxl

from targets import db as tdb

# Column aliases (lower-cased, spaces stripped) → canonical field.
_ALIASES = {
    "name": {"name", "leadname", "buyer", "buyername", "customer", "customername", "contactname", "fullname"},
    "phone": {"phone", "mobile", "number", "phonenumber", "mobilenumber", "contact", "contactnumber", "whatsapp"},
    "callnumber": {"callnumber", "callno", "aicall", "aicallnumber", "mobileforcall",
                   "dialnumber", "callmobile", "phoneforcall"},
    "config": {"config", "configuration", "unit", "bhk", "type", "requirement", "flat"},
    "budget": {"budget", "price", "amount"},
    "sector": {"sector", "product", "configbudget"},
    "status": {"status", "stage", "funnelstage"},
    "score": {"score", "rating"},
    "city": {"city", "location", "area"},
    "notes": {"notes", "note", "interaction", "interactions", "context", "remarks",
              "comments", "comment", "history", "conversation"},
    "email": {"email", "emailid", "mail"},
}
_STATUS_MAP = {
    "new": "new", "enquiry": "new", "fresh": "new",
    "contacted": "contacted", "called": "contacted", "qualified": "contacted",
    "meeting": "meeting", "sitevisit": "meeting", "visit": "meeting",
    "negotiation": "poc", "poc": "poc", "booking": "poc",
    "won": "won", "booked": "won", "closed": "won",
    "lost": "lost", "dropped": "lost",
}


def _canon(header: str) -> str:
    key = "".join(str(header or "").lower().split())
    for field, names in _ALIASES.items():
        if key in names:
            return field
    return ""


def parse_leads_xlsx(file_bytes: bytes, project_id: str,
                     owner_role: str = "presales") -> Dict[str, Any]:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        return {"ok": False, "error": f"Couldn't read the file: {str(e)[:120]}"}
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"ok": False, "error": "The sheet is empty."}

    # Header → column index map.
    header = rows[0]
    colmap: Dict[str, int] = {}
    for i, h in enumerate(header):
        f = _canon(h)
        if f and f not in colmap:
            colmap[f] = i
    if "name" not in colmap:
        return {"ok": False, "error": "No 'Name' column found. Add a Name column."}

    def cell(row, field):
        i = colmap.get(field)
        return (str(row[i]).strip() if i is not None and i < len(row) and row[i] is not None else "")

    created: List[str] = []
    skipped = 0
    for row in rows[1:]:
        if not row or not cell(row, "name"):
            skipped += 1
            continue
        name = cell(row, "name")
        phone = cell(row, "phone")
        config = cell(row, "config") or cell(row, "sector")
        budget = cell(row, "budget")
        sector = (config + (" · " + budget if budget else "")) if config else budget
        status = _STATUS_MAP.get("".join(cell(row, "status").lower().split()), "new")
        call_num = cell(row, "callnumber") or phone
        lead_id = tdb.create_lead(name, project_id, owner_role=owner_role,
                                  phone=call_num or phone, sector=sector, status=status,
                                  dm_email=cell(row, "email"))
        # Register the AI-call number so the Call button works right after import.
        if call_num:
            tdb.add_lead_number(lead_id, call_num, "AI call")
        # Carry the context in as the lead's starting memory.
        notes = cell(row, "notes")
        if notes:
            tdb.add_note(lead_id, "note", f"[imported] {notes[:400]}")
        score = cell(row, "score")
        if score:
            tdb.add_note(lead_id, "insight", f"[imported] prior score: {score}")
        city = cell(row, "city")
        if city:
            tdb.add_note(lead_id, "note", f"[imported] location: {city}")
        # Run intent scoring over any imported interaction text.
        if notes:
            try:
                from targets import intent_scoring
                intent_scoring.apply(lead_id, notes, "note", tdb.get_company(lead_id) or {})
            except Exception:  # noqa: BLE001
                pass
        created.append(lead_id)
    return {"ok": True, "created": len(created), "skipped": skipped, "ids": created}


def template_xlsx() -> bytes:
    """A sample sheet the user can fill + upload."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"
    ws.append(["Name", "Phone", "Call Number", "Config", "Budget", "Status", "Score", "City", "Notes"])
    ws.append(["Vinita Sharma", "+910000000091", "+910000000091", "3 BHK", "4.5 Cr", "Contacted", "80",
               "Gurgaon", "Called, wants site visit this month, ready to book"])
    ws.append(["Rahul Gupta", "+910000000092", "+910000000092", "4 BHK", "6.2 Cr", "New", "",
               "Delhi", "Website enquiry, comparing with DLF"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
